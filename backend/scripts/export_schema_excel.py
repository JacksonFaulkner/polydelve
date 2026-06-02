"""Export DuckDB tables to Excel for schema review."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import duckdb
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
DB_PATH = ROOT / "action_odds.duckdb"
OUTPUT_DIR = ROOT / "outputs"

# ── palette ──────────────────────────────────────────────────────────────────
DARK_BG    = "1F2937"
MID_BG     = "374151"
LIGHT_BG   = "4B5563"
ACCENT     = "3B82F6"
ACCENT2    = "10B981"
WARN       = "F59E0B"
HEADER_FG  = "FFFFFF"
ALT_ROW    = "F9FAFB"
BORDER_CLR = "D1D5DB"
SHEET_TITLE_FG = "F3F4F6"

# ── helpers ───────────────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=None, size=11, italic=False) -> Font:
    kwargs: dict = {"bold": bold, "size": size, "italic": italic}
    if color:
        kwargs["color"] = color
    return Font(**kwargs)

def _border(color=BORDER_CLR) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _left(wrap=False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _col_letter(n: int) -> str:
    return get_column_letter(n)

def _autofit(ws, min_width=8, max_width=60):
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[_col_letter(col_cells[0].column)].width = min(
            max(length + 2, min_width), max_width
        )

def _freeze(ws, row=2, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)


# ── sheet builders ─────────────────────────────────────────────────────────

def build_cover_sheet(wb: Workbook, tables_meta: list[dict]):
    ws = wb.create_sheet("Overview", 0)
    ws.sheet_view.showGridLines = False

    # title block
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Action Odds — DuckDB Schema Overview"
    c.font = _font(bold=True, color=HEADER_FG, size=16)
    c.fill = _fill(DARK_BG)
    c.alignment = _center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  DB: {DB_PATH.name}"
    c.font = _font(color="9CA3AF", size=10, italic=True)
    c.fill = _fill(MID_BG)
    c.alignment = _center()
    ws.row_dimensions[2].height = 20

    # spacer
    ws.row_dimensions[3].height = 10

    # header row
    headers = ["Table", "Rows", "Columns", "PK / Key Cols", "Notable Columns", "Sheet"]
    widths   = [22, 10, 10, 30, 42, 18]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = _font(bold=True, color=HEADER_FG, size=11)
        c.fill = _fill(ACCENT)
        c.alignment = _center()
        c.border = _border("2F4ACF")
        ws.column_dimensions[_col_letter(i)].width = w
    ws.row_dimensions[4].height = 22

    for row_idx, meta in enumerate(tables_meta, start=5):
        alt = row_idx % 2 == 0
        row_fill = _fill(ALT_ROW) if alt else _fill("FFFFFF")
        values = [
            meta["table"],
            meta["row_count"],
            meta["col_count"],
            meta["pk_cols"],
            meta["notable"],
            meta["sheet_name"],
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill = row_fill
            c.border = _border()
            c.alignment = _left(wrap=(col_idx in (4, 5)))
            if col_idx == 1:
                c.font = _font(bold=True, size=11)
            elif col_idx == 2:
                c.alignment = _center()
                if isinstance(val, int) and val > 0:
                    c.font = _font(bold=True, color="065F46", size=11)
            elif col_idx == 6:
                c.font = _font(bold=False, color=ACCENT, size=11, italic=True)
            else:
                c.font = _font(size=10)
        ws.row_dimensions[row_idx].height = 18

    # total row
    total_row = len(tables_meta) + 5
    ws.merge_cells(f"A{total_row}:E{total_row}")
    c = ws.cell(row=total_row, column=1, value=f"Total: {len(tables_meta)} tables")
    c.font = _font(bold=True, color=HEADER_FG, size=11)
    c.fill = _fill(MID_BG)
    c.alignment = _left()
    c = ws.cell(row=total_row, column=6, value=sum(m["row_count"] for m in tables_meta))
    c.font = _font(bold=True, color=HEADER_FG, size=11)
    c.fill = _fill(MID_BG)
    c.alignment = _center()


def build_schema_sheet(wb: Workbook, table: str, con: duckdb.DuckDBPyConnection):
    """One sheet per table: full DESCRIBE output with sample stats."""
    sheet_name = f"schema_{table}"[:31]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # banner
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"  Table: {table}"
    c.font = _font(bold=True, color=HEADER_FG, size=13)
    c.fill = _fill(DARK_BG)
    c.alignment = _left()
    ws.row_dimensions[1].height = 30

    describe_rows = con.execute(f"DESCRIBE {table}").fetchall()
    row_count = con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"  {row_count:,} rows  ·  {len(describe_rows)} columns"
    c.font = _font(color="9CA3AF", size=10, italic=True)
    c.fill = _fill(MID_BG)
    c.alignment = _left()
    ws.row_dimensions[2].height = 18

    # column headers
    col_headers = ["#", "Column Name", "Data Type", "Nullable", "Key", "Default", "Min (sample)", "Max (sample)"]
    col_widths   = [5,  24,            22,          10,         10,    16,        22,              22]
    for i, (h, w) in enumerate(zip(col_headers, col_widths), start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = _font(bold=True, color=HEADER_FG, size=10)
        c.fill = _fill(ACCENT)
        c.alignment = _center()
        c.border = _border()
        ws.column_dimensions[_col_letter(i)].width = w
    ws.row_dimensions[3].height = 20
    _freeze(ws, row=4)

    # describe cols: column_name, column_type, null, key, default, extra
    for row_idx, row in enumerate(describe_rows, start=4):
        col_name  = row[0]
        col_type  = row[1]
        nullable  = row[2]
        key       = row[3] if len(row) > 3 else ""
        default   = row[4] if len(row) > 4 else ""

        alt = row_idx % 2 == 0
        bg = _fill(ALT_ROW) if alt else _fill("FFFFFF")

        is_array = "[]" in col_type
        is_vec   = col_type.startswith("FLOAT[")
        is_key   = key and key.strip()

        mn = mx = ""
        if not is_vec and not is_array and row_count > 0:
            try:
                r = con.execute(
                    f"SELECT MIN(\"{col_name}\"), MAX(\"{col_name}\") FROM {table}"
                ).fetchone()
                mn = str(r[0]) if r[0] is not None else "NULL"
                mx = str(r[1]) if r[1] is not None else "NULL"
                if len(mn) > 40:
                    mn = mn[:37] + "..."
                if len(mx) > 40:
                    mx = mx[:37] + "..."
            except Exception:
                mn = mx = "—"
        elif is_vec:
            mn = mx = f"float[{col_type.split('[')[1].rstrip(']')}]"
        elif is_array and row_count > 0:
            try:
                r = con.execute(
                    f"SELECT MIN(len(\"{col_name}\")), MAX(len(\"{col_name}\")) FROM {table}"
                ).fetchone()
                mn = f"len≥{r[0]}"
                mx = f"len≤{r[1]}"
            except Exception:
                mn = mx = "—"

        cells_data = [
            row_idx - 3,
            col_name,
            col_type,
            nullable or "",
            key or "",
            default or "",
            mn,
            mx,
        ]
        for col_idx, val in enumerate(cells_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill = bg
            c.border = _border()
            c.alignment = _center() if col_idx in (1, 4, 5) else _left()
            if col_idx == 2:
                c.font = _font(bold=True, size=10, color=DARK_BG)
                if is_key:
                    c.font = _font(bold=True, size=10, color="7C3AED")
            elif col_idx == 3:
                color = "065F46" if "VARCHAR" in col_type else (
                    "1E40AF" if "INT" in col_type or "FLOAT" in col_type else (
                        "78350F" if "TIMESTAMP" in col_type else (
                            "6B21A8" if is_vec else "374151"
                        )
                    )
                )
                c.font = _font(size=10, color=color, italic=is_vec)
            elif col_idx == 5 and is_key:
                c.fill = _fill("EDE9FE")
                c.font = _font(bold=True, size=10, color="7C3AED")
            else:
                c.font = _font(size=9)
        ws.row_dimensions[row_idx].height = 16

    return sheet_name


def build_data_sheet(
    wb: Workbook,
    table: str,
    con: duckdb.DuckDBPyConnection,
    max_rows: int = 200,
    skip_vec_cols: bool = True,
):
    """Sample data sheet for a table — skips float vector columns."""
    sheet_name = f"data_{table}"[:31]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    describe_rows = con.execute(f"DESCRIBE {table}").fetchall()

    # filter out vector embedding columns
    cols = []
    for r in describe_rows:
        col_name = r[0]
        col_type = r[1]
        if skip_vec_cols and col_type.startswith("FLOAT["):
            continue
        cols.append((col_name, col_type))

    if not cols:
        ws["A1"] = "No displayable columns (all are float vectors)."
        return sheet_name

    col_select = ", ".join(f'"{c[0]}"' for c in cols)
    row_count  = con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
    fetch_n    = min(max_rows, row_count)

    # banner
    ws.merge_cells(f"A1:{_col_letter(len(cols))}1")
    c = ws["A1"]
    c.value = f"  Data: {table}  (showing {fetch_n:,} / {row_count:,} rows)"
    c.font  = _font(bold=True, color=HEADER_FG, size=12)
    c.fill  = _fill(DARK_BG)
    c.alignment = _left()
    ws.row_dimensions[1].height = 28

    # col headers
    for i, (col_name, col_type) in enumerate(cols, start=1):
        c = ws.cell(row=2, column=i, value=col_name)
        c.font = _font(bold=True, color=HEADER_FG, size=10)
        c.fill = _fill(ACCENT2)
        c.alignment = _center()
        c.border = _border()
    ws.row_dimensions[2].height = 20
    _freeze(ws, row=3)

    # type hint row
    for i, (col_name, col_type) in enumerate(cols, start=1):
        c = ws.cell(row=3, column=i, value=col_type)
        c.font = _font(color="6B7280", size=8, italic=True)
        c.fill = _fill("E5E7EB")
        c.alignment = _center()
        c.border = _border()
    ws.row_dimensions[3].height = 14

    # data rows
    if fetch_n > 0:
        rows = con.execute(
            f"SELECT {col_select} FROM {table} LIMIT {fetch_n}"
        ).fetchall()

        for row_idx, row in enumerate(rows, start=4):
            alt = row_idx % 2 == 0
            bg  = _fill(ALT_ROW) if alt else _fill("FFFFFF")
            for col_idx, val in enumerate(row, start=1):
                col_type = cols[col_idx - 1][1]
                # flatten arrays to comma-joined string; strip tz from datetimes
                if isinstance(val, list):
                    display = ", ".join(str(v) for v in val) if val else ""
                elif val is None:
                    display = ""
                elif hasattr(val, "tzinfo") and val.tzinfo is not None:
                    display = val.replace(tzinfo=None)
                else:
                    display = val
                c = ws.cell(row=row_idx, column=col_idx, value=display)
                c.fill = bg
                c.border = _border()
                c.alignment = _left(wrap=False)
                c.font = _font(size=9)
            ws.row_dimensions[row_idx].height = 15

    _autofit(ws, max_width=50)
    return sheet_name


def build_cross_ref_sheet(wb: Workbook, con: duckdb.DuckDBPyConnection):
    """Cross-table join counts — shows how tables relate."""
    ws = wb.create_sheet("Cross-Ref")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Cross-Table Relationship Counts"
    c.font  = _font(bold=True, color=HEADER_FG, size=13)
    c.fill  = _fill(DARK_BG)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    queries = [
        ("news_packages → packages",
         "SELECT COUNT(*) FROM news_packages np "
         "JOIN packages p ON np.name = p.name AND np.ecosystem = p.ecosystem"),
        ("news_packages → news",
         "SELECT COUNT(*) FROM news_packages np JOIN news n ON np.news_id = n.id"),
        ("news: have company_labels",
         "SELECT COUNT(*) FROM news WHERE len(company_labels) > 0"),
        ("news: have sector_labels",
         "SELECT COUNT(*) FROM news WHERE len(sector_labels) > 0"),
        ("packages: have cve_ids",
         "SELECT COUNT(*) FROM packages WHERE len(cve_ids) > 0"),
        ("packages: have sectors",
         "SELECT COUNT(*) FROM packages WHERE len(sectors) > 0"),
        ("packages: has_mal_advisory=true",
         "SELECT COUNT(*) FROM packages WHERE has_mal_advisory = true"),
        ("packages: weekly_downloads > 0",
         "SELECT COUNT(*) FROM packages WHERE weekly_downloads > 0"),
        ("cve_history: unique packages",
         "SELECT COUNT(DISTINCT name || '/' || ecosystem) FROM cve_history"),
        ("cve_history: have cvss_vector",
         "SELECT COUNT(*) FROM cve_history WHERE cvss_vector IS NOT NULL"),
        ("news_duplicates: unique candidate_urls",
         "SELECT COUNT(DISTINCT candidate_url) FROM news_duplicates"),
        ("news: avg embed title NULL",
         "SELECT COUNT(*) FROM news WHERE embed_title IS NULL"),
        ("packages: pypi ecosystem",
         "SELECT COUNT(*) FROM packages WHERE ecosystem = 'PyPI'"),
        ("packages: npm ecosystem",
         "SELECT COUNT(*) FROM packages WHERE ecosystem = 'npm'"),
        ("cve_history: HIGH or CRITICAL severity",
         "SELECT COUNT(*) FROM cve_history WHERE severity IN ('HIGH','CRITICAL')"),
    ]

    headers = ["Query / Metric", "Count"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _font(bold=True, color=HEADER_FG, size=11)
        c.fill = _fill(ACCENT)
        c.alignment = _center()
        c.border = _border()
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 14
    ws.row_dimensions[2].height = 20

    for row_idx, (label, sql) in enumerate(queries, start=3):
        alt = row_idx % 2 == 0
        bg  = _fill(ALT_ROW) if alt else _fill("FFFFFF")
        try:
            result = con.execute(sql).fetchone()[0]
        except Exception as e:
            result = f"ERR: {e}"

        c = ws.cell(row=row_idx, column=1, value=label)
        c.fill = bg
        c.border = _border()
        c.alignment = _left()
        c.font = _font(size=10)

        c = ws.cell(row=row_idx, column=2, value=result)
        c.fill = bg
        c.border = _border()
        c.alignment = _center()
        c.font = _font(bold=True, size=10,
                       color="065F46" if isinstance(result, int) and result > 0 else "374151")
        ws.row_dimensions[row_idx].height = 16


def build_packages_sectors_sheet(wb: Workbook, con: duckdb.DuckDBPyConnection):
    """Breakdown of package sectors."""
    ws = wb.create_sheet("Package Sectors")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "Package Sector Breakdown"
    c.font  = _font(bold=True, color=HEADER_FG, size=13)
    c.fill  = _fill(DARK_BG)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    # unnest sectors and count
    sql = """
        SELECT sector, COUNT(*) as pkg_count
        FROM (
            SELECT unnest(sectors) as sector FROM packages WHERE sectors IS NOT NULL
        )
        GROUP BY sector
        ORDER BY pkg_count DESC
    """
    rows = con.execute(sql).fetchall()

    headers = ["Sector", "Package Count"]
    widths  = [36, 16]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _font(bold=True, color=HEADER_FG, size=11)
        c.fill = _fill(ACCENT2)
        c.alignment = _center()
        c.border = _border()
        ws.column_dimensions[_col_letter(i)].width = w
    ws.row_dimensions[2].height = 20

    if not rows:
        ws.cell(row=3, column=1, value="No sector data yet.")
    for row_idx, (sector, count) in enumerate(rows, start=3):
        alt = row_idx % 2 == 0
        bg  = _fill(ALT_ROW) if alt else _fill("FFFFFF")
        c = ws.cell(row=row_idx, column=1, value=sector)
        c.fill = bg; c.border = _border(); c.alignment = _left()
        c.font = _font(size=10)
        c = ws.cell(row=row_idx, column=2, value=count)
        c.fill = bg; c.border = _border(); c.alignment = _center()
        c.font = _font(bold=True, size=10, color="1E40AF")
        ws.row_dimensions[row_idx].height = 16


def build_cve_severity_sheet(wb: Workbook, con: duckdb.DuckDBPyConnection):
    """CVE history severity distribution."""
    ws = wb.create_sheet("CVE Severity")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "CVE History — Severity Distribution"
    c.font  = _font(bold=True, color=HEADER_FG, size=13)
    c.fill  = _fill(DARK_BG)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    sev_sql = """
        SELECT COALESCE(severity, 'UNKNOWN') as sev, COUNT(*) as cnt
        FROM cve_history GROUP BY sev ORDER BY cnt DESC
    """
    eco_sql = """
        SELECT ecosystem, COUNT(*) as cnt
        FROM cve_history GROUP BY ecosystem ORDER BY cnt DESC
    """
    sev_rows = con.execute(sev_sql).fetchall()
    eco_rows = con.execute(eco_sql).fetchall()

    # severity table
    for i, h in enumerate(["Severity", "Count"], start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _font(bold=True, color=HEADER_FG, size=11)
        c.fill = _fill(ACCENT)
        c.alignment = _center()
        c.border = _border()
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12

    sev_colors = {
        "CRITICAL": "7F1D1D", "HIGH": "B45309", "MEDIUM": "1E3A5F",
        "LOW": "065F46", "UNKNOWN": "4B5563",
    }
    for row_idx, (sev, cnt) in enumerate(sev_rows, start=3):
        bg  = _fill(ALT_ROW if row_idx % 2 == 0 else "FFFFFF")
        c = ws.cell(row=row_idx, column=1, value=sev)
        c.fill = bg; c.border = _border(); c.alignment = _left()
        c.font = _font(bold=True, size=10, color=sev_colors.get(sev, "374151"))
        c = ws.cell(row=row_idx, column=2, value=cnt)
        c.fill = bg; c.border = _border(); c.alignment = _center()
        c.font = _font(size=10)
        ws.row_dimensions[row_idx].height = 16

    # ecosystem table — offset to col D
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    for i, h in enumerate(["Ecosystem", "Count"], start=4):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _font(bold=True, color=HEADER_FG, size=11)
        c.fill = _fill(WARN)
        c.alignment = _center()
        c.border = _border()
    for row_idx, (eco, cnt) in enumerate(eco_rows, start=3):
        bg = _fill(ALT_ROW if row_idx % 2 == 0 else "FFFFFF")
        c = ws.cell(row=row_idx, column=4, value=eco)
        c.fill = bg; c.border = _border(); c.alignment = _left()
        c.font = _font(size=10)
        c = ws.cell(row=row_idx, column=5, value=cnt)
        c.fill = bg; c.border = _border(); c.alignment = _center()
        c.font = _font(size=10)
        ws.row_dimensions[row_idx].height = 16


def build_news_labels_sheet(wb: Workbook, con: duckdb.DuckDBPyConnection):
    """News company label and sector label frequency."""
    ws = wb.create_sheet("News Labels")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "News — Company & Sector Label Frequency"
    c.font  = _font(bold=True, color=HEADER_FG, size=13)
    c.fill  = _fill(DARK_BG)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    company_sql = """
        SELECT label, COUNT(*) as cnt
        FROM (SELECT unnest(company_labels) as label FROM news WHERE company_labels IS NOT NULL)
        GROUP BY label ORDER BY cnt DESC LIMIT 30
    """
    sector_sql = """
        SELECT label, COUNT(*) as cnt
        FROM (SELECT unnest(sector_labels) as label FROM news WHERE sector_labels IS NOT NULL)
        GROUP BY label ORDER BY cnt DESC LIMIT 30
    """
    company_rows = con.execute(company_sql).fetchall()
    sector_rows  = con.execute(sector_sql).fetchall()

    # company labels cols A-B
    for i, h in enumerate(["Company Label", "News Count"], start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _font(bold=True, color=HEADER_FG, size=11)
        c.fill = _fill(ACCENT)
        c.alignment = _center()
        c.border = _border()
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14

    for row_idx, (label, cnt) in enumerate(company_rows, start=3):
        bg = _fill(ALT_ROW if row_idx % 2 == 0 else "FFFFFF")
        c = ws.cell(row=row_idx, column=1, value=label)
        c.fill = bg; c.border = _border(); c.alignment = _left()
        c.font = _font(bold=True, size=10)
        c = ws.cell(row=row_idx, column=2, value=cnt)
        c.fill = bg; c.border = _border(); c.alignment = _center()
        c.font = _font(size=10)
        ws.row_dimensions[row_idx].height = 16

    # sector labels cols D-E
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 14
    for i, h in enumerate(["Sector Label", "News Count"], start=4):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _font(bold=True, color=HEADER_FG, size=11)
        c.fill = _fill(ACCENT2)
        c.alignment = _center()
        c.border = _border()
    for row_idx, (label, cnt) in enumerate(sector_rows, start=3):
        bg = _fill(ALT_ROW if row_idx % 2 == 0 else "FFFFFF")
        c = ws.cell(row=row_idx, column=4, value=label)
        c.fill = bg; c.border = _border(); c.alignment = _left()
        c.font = _font(size=10)
        c = ws.cell(row=row_idx, column=5, value=cnt)
        c.fill = bg; c.border = _border(); c.alignment = _center()
        c.font = _font(size=10)
        ws.row_dimensions[row_idx].height = 16


# ── meta helpers ──────────────────────────────────────────────────────────────

def _build_meta(table: str, con: duckdb.DuckDBPyConnection, sheet_name: str) -> dict:
    row_count = con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
    describe  = con.execute(f"DESCRIBE {table}").fetchall()
    pk_cols   = [r[0] for r in describe if r[3] and "PRI" in str(r[3]).upper()]
    if not pk_cols:
        pk_cols = [r[0] for r in describe if r[3]]
    notable   = [r[0] for r in describe if r[1].startswith("FLOAT[") or "[]" in r[1]]
    return {
        "table":      table,
        "row_count":  row_count,
        "col_count":  len(describe),
        "pk_cols":    ", ".join(pk_cols) or "—",
        "notable":    ", ".join(notable) or "—",
        "sheet_name": sheet_name,
    }


# ── main ───────────────────────────────────────────────────────────────────────

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUTPUT_DIR / f"schema_overview_{stamp}.xlsx"

    print(f"Connecting to {DB_PATH} ...")
    con = duckdb.connect(str(DB_PATH), read_only=True)

    tables = [r[0] for r in con.execute("SHOW TABLES").fetchall()]
    print(f"Found {len(tables)} tables: {tables}")

    wb = Workbook()
    # remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 1. per-table schema + data sheets
    tables_meta = []
    for table in tables:
        print(f"  Building schema sheet: {table}")
        sn = build_schema_sheet(wb, table, con)
        meta = _build_meta(table, con, sn)
        tables_meta.append(meta)
        print(f"  Building data sheet:   {table} ({meta['row_count']} rows)")
        build_data_sheet(wb, table, con)

    # 2. overview (prepend)
    print("Building overview sheet ...")
    build_cover_sheet(wb, tables_meta)
    wb.move_sheet("Overview", offset=-len(wb.sheetnames) + 1)

    # 3. analysis sheets
    print("Building cross-ref sheet ...")
    build_cross_ref_sheet(wb, con)

    print("Building package sectors sheet ...")
    build_packages_sectors_sheet(wb, con)

    print("Building CVE severity sheet ...")
    build_cve_severity_sheet(wb, con)

    print("Building news labels sheet ...")
    build_news_labels_sheet(wb, con)

    con.close()

    wb.save(str(out_path))
    print(f"\nSaved: {out_path}")
    return out_path


if __name__ == "__main__":
    path = main()
    print(f"Done. Open: {path}")
